import asyncio
import os.path

from openpyxl.drawing.image import Image
from openpyxl.workbook import Workbook

from src.website.utils.excel_creator.schemas import DataClass

current_directory = os.path.abspath(__file__)
directory_dataset = os.path.join(current_directory, os.pardir, os.pardir, os.pardir, os.pardir, "resources", "yoga-82-dataset", "test")
excel_file = os.path.join(current_directory, os.pardir, "excel.xlsx")
wb = Workbook()
sheet = wb.active



async def excel_creator():
    list_classes: list[DataClass] = []
    list_directory_class = os.listdir(directory_dataset)

    for directory_class in list_directory_class:
        source_name = directory_class
        russian_name = ""
        image = Image(os.path.join(directory_dataset, directory_class, os.listdir(os.path.join(directory_dataset, directory_class))[0]))

        max_size = 300
        if image.width > max_size or image.height > max_size:
            scale = max_size / max(image.width, image.height)
            image.width *= scale
            image.height *= scale

        list_classes.append(DataClass(source_name=source_name, russian_name=russian_name, image=image))

    for index, class_item in enumerate(list_classes):
        sheet.cell(row=index+1, column=1).value = class_item.source_name
        sheet.cell(row=index+1, column=2).value = class_item.russian_name
        sheet.add_image(class_item.image, f"C{index+1}")


    wb.save('excel.xlsx')

if __name__ == "__main__":
    asyncio.run(excel_creator())